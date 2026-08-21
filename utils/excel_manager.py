import csv
import json
import os
from pathlib import Path
from datetime import date, datetime
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _get_data_dir() -> Path:
    try:
        from utils.local_settings import get_data_dir
        d = get_data_dir()
    except Exception:
        d = None
    if d is None:
        d = Path.home() / "GansikdangData"
    d.mkdir(parents=True, exist_ok=True)
    return d

SHEET_DAILY = "일별매출"
SHEET_EXPENSE = "매입내역"
SHEET_TAX = "세금내역"
SHEET_SALARY = "급여내역"

DAILY_HEADERS = ["날짜", "카드(ft)", "NAV Cash(ft)", "Invoice(ft)", "현금(ft)", "계좌이체합계(ft)", "계좌이체상세(JSON)", "카드건수", "NAV건수"]
EXPENSE_HEADERS = ["날짜", "구매처", "합계(ft)", "면세(ft)", "5%(ft)", "18%(ft)", "27%(ft)", "결제방법", "카테고리", "비고"]
TAX_HEADERS = ["날짜", "세금종류", "금액(ft)"]
SALARY_HEADERS = ["날짜", "직원명", "직급", "구분", "금액(ft)"]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _get_filepath(year: int, month: int) -> Path:
    return _get_data_dir() / f"{year:04d}_{month:02d}.xlsx"


def _style_header(ws, headers: list) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 20


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val.encode("euc-kr", errors="replace")))
            except Exception:
                max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _ensure_sheet(wb: Workbook, sheet_name: str, headers: list) -> openpyxl.worksheet.worksheet.Worksheet:
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    _style_header(ws, headers)
    return ws


def _load_or_create_wb(filepath: Path) -> Workbook:
    if filepath.exists():
        return load_workbook(filepath)
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def _find_row_by_date(ws, date_str: str) -> int | None:
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell_val = row[0].value
        if cell_val is not None and str(cell_val)[:10] == date_str:
            return row[0].row
    return None


def _style_data_rows(ws) -> None:
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if i % 2 == 0 else None
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill


# ─────────────────────────────────────────────
# 일별 매출
# ─────────────────────────────────────────────

def save_daily_sales(year: int, month: int, date_str: str, data: dict) -> None:
    """
    data keys: card, nav_cash, invoice, cash, transfer_total, transfer_detail (list of {name, amount})
    """
    fp = _get_filepath(year, month)
    wb = _load_or_create_wb(fp)
    ws = _ensure_sheet(wb, SHEET_DAILY, DAILY_HEADERS)

    transfer_json = json.dumps(data.get("transfer_detail", []), ensure_ascii=False)
    row_data = [
        date_str,
        data.get("card", 0),
        data.get("nav_cash", 0),
        data.get("invoice", 0),
        data.get("cash", 0),
        data.get("transfer_total", 0),
        transfer_json,
        data.get("card_count", 0),
        data.get("nav_count", 0),
    ]

    existing_row = _find_row_by_date(ws, date_str)
    if existing_row:
        for col, val in enumerate(row_data, 1):
            ws.cell(row=existing_row, column=col, value=val)
    else:
        ws.append(row_data)

    _style_data_rows(ws)
    _auto_width(ws)
    wb.save(fp)


def load_daily_sales(year: int, month: int) -> list[dict]:
    fp = _get_filepath(year, month)
    if not fp.exists():
        return []
    wb = load_workbook(fp)
    if SHEET_DAILY not in wb.sheetnames:
        return []
    ws = wb[SHEET_DAILY]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        try:
            transfer_detail = json.loads(row[6]) if row[6] else []
        except Exception:
            transfer_detail = []
        result.append({
            "date": str(row[0])[:10],
            "card": row[1] or 0,
            "nav_cash": row[2] or 0,
            "invoice": row[3] or 0,
            "cash": row[4] or 0,
            "transfer_total": row[5] or 0,
            "transfer_detail": transfer_detail,
            "card_count": (int(row[7]) if len(row) > 7 and row[7] else 0),
            "nav_count": (int(row[8]) if len(row) > 8 and row[8] else 0),
        })
    result.sort(key=lambda x: x["date"])
    return result


# ─────────────────────────────────────────────
# 매입내역
# ─────────────────────────────────────────────

def save_expense(year: int, month: int, entries: list[dict]) -> None:
    fp = _get_filepath(year, month)
    wb = _load_or_create_wb(fp)
    ws = _ensure_sheet(wb, SHEET_EXPENSE, EXPENSE_HEADERS)

    for e in entries:
        ws.append([
            e.get("date", ""),
            e.get("vendor", ""),
            e.get("amount_ft", 0),
            e.get("amount_none", 0),
            e.get("amount_5", 0),
            e.get("amount_18", 0),
            e.get("amount_27", 0),
            e.get("payment_method", "카드"),
            e.get("category", "식재료 및 판관비"),
            e.get("note", ""),
        ])

    _style_data_rows(ws)
    _auto_width(ws)
    wb.save(fp)


def replace_expenses(year: int, month: int, entries: list[dict]) -> None:
    """월 전체 매입내역을 entries로 교체 (수정/삭제용)."""
    fp = _get_filepath(year, month)
    wb = _load_or_create_wb(fp)
    if SHEET_EXPENSE in wb.sheetnames:
        del wb[SHEET_EXPENSE]
    ws = _ensure_sheet(wb, SHEET_EXPENSE, EXPENSE_HEADERS)
    for e in entries:
        ws.append([
            e.get("date", ""),
            e.get("vendor", ""),
            e.get("amount_ft", 0),
            e.get("amount_none", 0),
            e.get("amount_5", 0),
            e.get("amount_18", 0),
            e.get("amount_27", 0),
            e.get("payment_method", "카드"),
            e.get("category", "식재료 및 판관비"),
            e.get("note", ""),
        ])
    _style_data_rows(ws)
    _auto_width(ws)
    wb.save(fp)


def sync_vendor_categories_in_expenses(category_map: dict) -> int:
    """업체명 → 분류 맵을 받아 모든 엑셀 매입 기록을 한 번에 동기화.
    category_map = {"업체명": "새분류", ...}
    이미 일치하는 레코드는 건드리지 않음.
    """
    updated = 0
    for fp in _get_data_dir().glob("*.xlsx"):
        parts = fp.stem.split("_")
        if len(parts) != 2:
            continue
        try:
            y, m = int(parts[0]), int(parts[1])
        except ValueError:
            continue
        records = load_expenses(y, m)
        changed = False
        for r in records:
            vendor = r.get("vendor", "")
            if vendor in category_map:
                new_cat = category_map[vendor]
                if new_cat and r.get("category") != new_cat:
                    r["category"] = new_cat
                    updated += 1
                    changed = True
        if changed:
            replace_expenses(y, m, records)
    return updated


def update_vendor_category_in_expenses(vendor_name: str, new_category: str) -> int:
    """단일 업체 분류 업데이트 (legacy용)."""
    return sync_vendor_categories_in_expenses({vendor_name: new_category})


def rename_vendor_in_expenses(old_name: str, new_name: str) -> int:
    """모든 매입 기록에서 업체명 변경. 변경된 건수 반환."""
    count = 0
    for fp in _get_data_dir().glob("*.xlsx"):
        parts = fp.stem.split("_")
        if len(parts) != 2:
            continue
        try:
            y, m = int(parts[0]), int(parts[1])
        except ValueError:
            continue
        records = load_expenses(y, m)
        changed = False
        for r in records:
            if r.get("vendor") == old_name:
                r["vendor"] = new_name
                count += 1
                changed = True
        if changed:
            replace_expenses(y, m, records)
    return count


def load_expenses(year: int, month: int) -> list[dict]:
    fp = _get_filepath(year, month)
    if not fp.exists():
        return []
    wb = load_workbook(fp)
    if SHEET_EXPENSE not in wb.sheetnames:
        return []
    ws = wb[SHEET_EXPENSE]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        result.append({
            "date": str(row[0])[:10],
            "vendor": row[1] or "",
            "amount_ft": row[2] or 0,
            "amount_none": float(row[3] or 0),
            "amount_5":    float(row[4] or 0),
            "amount_18":   float(row[5] or 0),
            "amount_27":   float(row[6] or 0),
            "payment_method": (row[7] if len(row) > 7 and row[7] else "카드"),
            "category":       (row[8] if len(row) > 8 and row[8] else "식재료 및 판관비"),
            "note":           (row[9] if len(row) > 9 and row[9] else ""),
        })
    result.sort(key=lambda x: x["date"])
    return result


# ─────────────────────────────────────────────
# 세금내역
# ─────────────────────────────────────────────

def save_tax(year: int, month: int, entries: list[dict]) -> None:
    """entries: list of {date, tax_name, amount_ft}"""
    fp = _get_filepath(year, month)
    wb = _load_or_create_wb(fp)
    ws = _ensure_sheet(wb, SHEET_TAX, TAX_HEADERS)

    for e in entries:
        ws.append([
            e.get("date", ""),
            e.get("tax_name", ""),
            e.get("amount_ft", 0),
        ])

    _style_data_rows(ws)
    _auto_width(ws)
    wb.save(fp)


def load_taxes(year: int, month: int) -> list[dict]:
    fp = _get_filepath(year, month)
    if not fp.exists():
        return []
    wb = load_workbook(fp)
    if SHEET_TAX not in wb.sheetnames:
        return []
    ws = wb[SHEET_TAX]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        result.append({
            "date": str(row[0])[:10],
            "tax_name": row[1] or "",
            "amount_ft": row[2] or 0,
        })
    result.sort(key=lambda x: x["date"])
    return result


def replace_taxes(year: int, month: int, entries: list[dict]) -> None:
    """월 전체 세금내역을 entries로 교체 (삭제용)."""
    fp = _get_filepath(year, month)
    wb = _load_or_create_wb(fp)
    if SHEET_TAX in wb.sheetnames:
        del wb[SHEET_TAX]
    ws = _ensure_sheet(wb, SHEET_TAX, TAX_HEADERS)
    for e in entries:
        ws.append([e.get("date", ""), e.get("tax_name", ""), e.get("amount_ft", 0)])
    _style_data_rows(ws)
    _auto_width(ws)
    wb.save(fp)


# ─────────────────────────────────────────────
# 급여내역
# ─────────────────────────────────────────────

def save_salary(year: int, month: int, entry: dict) -> None:
    """entry: {date, name, position, category, amount_ft}"""
    fp = _get_filepath(year, month)
    wb = _load_or_create_wb(fp)
    ws = _ensure_sheet(wb, SHEET_SALARY, SALARY_HEADERS)

    ws.append([
        entry.get("date", ""),
        entry.get("name", ""),
        entry.get("position", ""),
        entry.get("category", ""),
        entry.get("amount_ft", 0),
    ])

    _style_data_rows(ws)
    _auto_width(ws)
    wb.save(fp)


def load_salaries(year: int, month: int) -> list[dict]:
    fp = _get_filepath(year, month)
    if not fp.exists():
        return []
    wb = load_workbook(fp)
    if SHEET_SALARY not in wb.sheetnames:
        return []
    ws = wb[SHEET_SALARY]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        result.append({
            "date": str(row[0])[:10],
            "name": row[1] or "",
            "position": row[2] or "",
            "category": row[3] or "",
            "amount_ft": row[4] or 0,
        })
    result.sort(key=lambda x: x["date"])
    return result


# ─────────────────────────────────────────────
# 직원 정보 (employees.json)
# ─────────────────────────────────────────────

def _employees_path() -> Path:
    return _get_data_dir() / "employees.json"


def _normalize_banks(raw: list) -> list[dict]:
    result = []
    for b in raw:
        if isinstance(b, str):
            result.append({"name": b, "currency": "ft", "account": ""})
        elif isinstance(b, dict):
            result.append({"name": b.get("name", ""), "currency": b.get("currency", "ft"),
                           "account": b.get("account", "")})
    return result


def load_employees() -> list[dict]:
    ep = _employees_path()
    if not ep.exists():
        return []
    with open(ep, "r", encoding="utf-8-sig") as f:
        data = json.load(f)
    for emp in data:
        if "banks" not in emp:
            old = emp.pop("bank", "")
            emp["banks"] = [old] if old else []
        emp["banks"] = _normalize_banks(emp["banks"])
    return data


def save_employees(employees: list[dict]) -> None:
    with open(_employees_path(), "w", encoding="utf-8") as f:
        json.dump(employees, f, ensure_ascii=False, indent=2)


def add_employee(name: str, position: str, banks: list[str], english_name: str = "",
                 official_account: str = "") -> None:
    employees = load_employees()
    employees.append({"name": name, "english_name": english_name, "official_account": official_account,
                      "position": position, "banks": banks})
    save_employees(employees)


def remove_employee(name: str) -> None:
    employees = [e for e in load_employees() if e["name"] != name]
    save_employees(employees)


def update_employee(name: str, updates: dict) -> None:
    employees = load_employees()
    for emp in employees:
        if emp["name"] == name:
            emp.update(updates)
            break
    save_employees(employees)


# ─────────────────────────────────────────────
# 선지급 잔액 (prepay_balance.json)
# ─────────────────────────────────────────────

def _prepay_path() -> Path:
    return _get_data_dir() / "prepay_balance.json"


def load_prepay() -> dict:
    """Returns {name: {"balance": float, "history": [...]}}"""
    pp = _prepay_path()
    if not pp.exists():
        return {}
    with open(pp, "r", encoding="utf-8-sig") as f:
        return json.load(f)


def save_prepay(data: dict) -> None:
    with open(_prepay_path(), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_prepay_balance(name: str) -> float:
    return load_prepay().get(name, {}).get("balance", 0.0)


def add_prepay(name: str, amount: float, method: str, date_str: str) -> None:
    data = load_prepay()
    if name not in data:
        data[name] = {"balance": 0.0, "history": []}
    data[name]["balance"] = round(data[name]["balance"] + amount, 2)
    data[name]["history"].append({"date": date_str, "method": method, "amount": amount})
    save_prepay(data)


def reset_prepay(name: str) -> None:
    data = load_prepay()
    if name in data:
        data[name]["balance"] = 0.0
    save_prepay(data)


def delete_prepay_entries_by_method(name: str, method: str, ym_prefix: str) -> None:
    """Remove prepay history entries matching method+month prefix and restore balance."""
    data = load_prepay()
    if name not in data:
        return
    old_hist = data[name].get("history", [])
    to_remove = [e for e in old_hist
                 if e.get("method") == method and e.get("date", "").startswith(ym_prefix)]
    if not to_remove:
        return
    removed_sum = sum(e.get("amount", 0) for e in to_remove)
    data[name]["history"] = [e for e in old_hist if e not in to_remove]
    data[name]["balance"] = round(data[name].get("balance", 0.0) - removed_sum, 2)
    save_prepay(data)


def transfer_prepay(from_name: str, to_name: str, amount: float, date_str: str) -> None:
    data = load_prepay()
    # 지급자: 잔액 차감
    if from_name not in data:
        data[from_name] = {"balance": 0.0, "history": []}
    data[from_name]["balance"] = round(data[from_name]["balance"] - amount, 2)
    data[from_name]["history"].append({
        "date": date_str, "method": f"{to_name}에게 지급", "amount": -amount,
    })
    # 수령자: 잔액 증가
    if to_name not in data:
        data[to_name] = {"balance": 0.0, "history": []}
    data[to_name]["balance"] = round(data[to_name]["balance"] + amount, 2)
    data[to_name]["history"].append({
        "date": date_str, "method": f"{from_name}으로부터 수령", "amount": amount,
    })
    save_prepay(data)
    # Excel 급여 기록 (회계용)
    y, m = int(date_str[:4]), int(date_str[5:7])
    save_salary(y, m, {
        "date": date_str, "name": to_name, "position": "",
        "category": "주급(직원간)", "amount_ft": amount,
    })


def get_prepay_history(name: str) -> list[dict]:
    return load_prepay().get(name, {}).get("history", [])


def delete_salary_records_by_category(year: int, month: int, name: str, category: str) -> int:
    fp = _get_filepath(year, month)
    if not fp.exists():
        return 0
    wb = load_workbook(fp)
    if SHEET_SALARY not in wb.sheetnames:
        return 0
    ws = wb[SHEET_SALARY]
    to_del = [row[0].row for row in ws.iter_rows(min_row=2)
              if (row[1].value or "") == name and (row[3].value or "") == category]
    for r in reversed(to_del):
        ws.delete_rows(r)
    if to_del:
        wb.save(fp)
    return len(to_del)


def delete_salary_record(year: int, month: int, name: str, date_str: str, category: str, amount: float) -> bool:
    fp = _get_filepath(year, month)
    if not fp.exists():
        return False
    wb = load_workbook(fp)
    if SHEET_SALARY not in wb.sheetnames:
        return False
    ws = wb[SHEET_SALARY]
    for row in ws.iter_rows(min_row=2):
        if (str(row[0].value or "")[:10] == date_str
                and (row[1].value or "") == name
                and (row[3].value or "") == category
                and abs(float(row[4].value or 0) - amount) < 0.01):
            ws.delete_rows(row[0].row)
            wb.save(fp)
            return True
    return False


def update_transfer_prepay(date_str: str, old_transfers: list[dict], new_transfers: list[dict]) -> None:
    """일일매출 계좌이체 변경 시 선지급 잔액 동기화."""
    data = load_prepay()

    # 기존 항목 제거
    for t in old_transfers:
        name = t.get("name", "")
        amt = t.get("amount", 0)
        if not name or name not in data:
            continue
        data[name]["balance"] = round(data[name]["balance"] - amt, 2)
        data[name]["history"] = [
            h for h in data[name]["history"]
            if not (h.get("date") == date_str and h.get("method", "").startswith("계좌이체") and h.get("amount") == amt)
        ]

    # 새 항목 추가
    for t in new_transfers:
        name = t.get("name", "")
        amt = t.get("amount", 0)
        bank = t.get("bank", "")
        if not name or amt <= 0:
            continue
        if name not in data:
            data[name] = {"balance": 0.0, "history": []}
        data[name]["balance"] = round(data[name]["balance"] + amt, 2)
        method = f"계좌이체({bank})" if bank else "계좌이체"
        data[name]["history"].append({"date": date_str, "method": method, "amount": amt})

    save_prepay(data)


# ─────────────────────────────────────────────
# 직원 급여 정보 (employee_salary.json)
# ─────────────────────────────────────────────

def _salary_info_path() -> Path:
    return _get_data_dir() / "employee_salary.json"


def load_salary_info() -> dict:
    sp = _salary_info_path()
    if not sp.exists():
        return {}
    with open(sp, "r", encoding="utf-8-sig") as f:
        return json.load(f)


def save_salary_info(data: dict) -> None:
    with open(_salary_info_path(), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_employee_salary(name: str) -> dict:
    defaults = {
        "official_salary": 0.0,
        "actual_salary": 0.0,
        "weekly_wage": 0.0,
        "part_tax": 0.0,
        "hourly_wage": 0.0,
    }
    saved = load_salary_info().get(name, {})
    return {**defaults, **saved}


def set_employee_salary(name: str, info: dict) -> None:
    data = load_salary_info()
    data[name] = info
    save_salary_info(data)


# ─────────────────────────────────────────────
# CSV 내보내기
# ─────────────────────────────────────────────

def export_month_csv(year: int, month: int, folder_path: str) -> list[str]:
    """Export all data for a given month to CSV files. Returns list of created file paths."""
    folder = Path(folder_path)
    folder.mkdir(parents=True, exist_ok=True)
    prefix = f"{year:04d}_{month:02d}"
    created = []

    sales = load_daily_sales(year, month)
    if sales:
        fp = folder / f"{prefix}_일별매출.csv"
        with open(fp, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "date", "card", "card_count", "nav_cash", "nav_count",
                "invoice", "cash", "transfer_total",
            ])
            writer.writeheader()
            for r in sales:
                writer.writerow({k: r.get(k, "") for k in writer.fieldnames})
        created.append(str(fp))

    expenses = load_expenses(year, month)
    if expenses:
        fp = folder / f"{prefix}_매입내역.csv"
        with open(fp, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "date", "category", "vendor", "amount_ft",
                "amount_none", "amount_5", "amount_18", "amount_27",
                "payment_method", "note",
            ])
            writer.writeheader()
            for r in expenses:
                writer.writerow({k: r.get(k, "") for k in writer.fieldnames})
        created.append(str(fp))

    salaries = load_salaries(year, month)
    if salaries:
        fp = folder / f"{prefix}_급여내역.csv"
        with open(fp, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=["date", "name", "position", "category", "amount_ft"])
            writer.writeheader()
            writer.writerows(salaries)
        created.append(str(fp))

    taxes = load_taxes(year, month)
    if taxes:
        fp = folder / f"{prefix}_세금내역.csv"
        with open(fp, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=["date", "tax_name", "amount_ft"])
            writer.writeheader()
            writer.writerows(taxes)
        created.append(str(fp))

    return created
